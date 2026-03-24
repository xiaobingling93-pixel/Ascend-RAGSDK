#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
-------------------------------------------------------------------------
This file is part of the RAGSDK project.
Copyright (c) 2025 Huawei Technologies Co.,Ltd.

RAGSDK is licensed under Mulan PSL v2.
You can use this software according to the terms and conditions of the Mulan PSL v2.
You may obtain a copy of Mulan PSL v2 at:

         http://license.coscl.org.cn/MulanPSL2

THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND,
EITHER EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT,
MERCHANTABILITY OR FIT FOR A PARTICULAR PURPOSE.
See the Mulan PSL v2 for more details.
-------------------------------------------------------------------------
"""

from datetime import datetime, timedelta
from typing import Iterator

import xlrd
from langchain_community.document_loaders.base import BaseLoader
from langchain_core.documents import Document
from loguru import logger
from openpyxl import load_workbook, Workbook
from openpyxl.cell import MergedCell

from mx_rag.document.loader.base_loader import BaseLoader as mxBaseLoader
from mx_rag.utils import file_check
from mx_rag.utils.common import validate_params, STR_TYPE_CHECK_TIP_1024, MAX_ROW_NUM, MAX_COL_NUM

OPENPYXL_EXTENSION = (".xlsx",)
XLRD_EXTENSION = (".xls",)


class ExcelLoader(BaseLoader, mxBaseLoader):
    @validate_params(
        line_sep=dict(validator=lambda x: isinstance(x, str) and 0 < len(x) <= 1024, message=STR_TYPE_CHECK_TIP_1024)
    )
    def __init__(self, file_path: str, line_sep: str = "**;"):
        super().__init__(file_path)
        self.line_sep = str(line_sep)

    @staticmethod
    def _exceltime_to_datetime(exceltime):
        """
        将excel储存的时间格式转换为可读的格式
        :param exceltime: (0,1)区间的浮点数，表示时间在一天中的位置
        :return: 以 时：分 的格式返回字符串
        """
        base_date = datetime(1899, 12, 30)
        time = base_date + timedelta(exceltime)
        return time.strftime("%H:%M")

    @staticmethod
    def _get_xlsx_blank_rows_and_cols(worksheet):
        """
        功能：获取所有空行、空列的索引
        """
        row_count = 0
        col_count = 0

        null_rows = {}
        for row in worksheet.iter_rows(values_only=True):
            row_count += 1
            if all(not cell for cell in row):
                null_rows[row_count] = True

        blank_cols = {}
        for col in worksheet.iter_cols(values_only=True):
            col_count += 1
            if all(not cell for cell in col):
                blank_cols[col_count] = True

        return null_rows, blank_cols

    @staticmethod
    def _get_xls_blank_rows_and_cols(worksheet):
        """
        功能：获取所有空行、空列的索引
        """
        null_rows = {}
        blank_cols = {}

        for i in range(worksheet.nrows):
            row = worksheet.row_values(i)
            if all(not cell for cell in row):
                null_rows[i] = True

        for i in range(worksheet.ncols):
            col = worksheet.col_values(i)
            if all(not cell for cell in col):
                blank_cols[i] = True

        return null_rows, blank_cols

    @staticmethod
    def _parse_xlsx_cell(sheet: Workbook, row: int, col: int):
        """
        功能：读取xlsx cell值，如果是合并项，使用左上cell值替代
        """
        cell = sheet.cell(row=row, column=col)
        if not isinstance(cell, MergedCell):
            return cell.value

        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # return the left top cell
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                return cell.value

        return cell.value

    @staticmethod
    def _parse_xls_cell(sheet, row: int, col: int):
        """
        功能：读取xls cell值，如果是合并项，使用左上cell值替代
        """
        cell = sheet.cell(rowx=row, colx=col)
        if cell.value:
            return cell.value

        for crange in sheet.merged_cells:
            rlo, rhi, clo, chi = crange
            if rlo <= row < rhi and clo <= col < chi:
                return sheet.cell(rowx=rlo, colx=clo).value

        return cell.value

    @staticmethod
    def _get_xlsx_first_not_blank_row_and_col(ws):
        """
        功能：读取xlsx 非空白首行、首列索引
        """
        first_row = 1
        first_col = 1
        for row in ws.iter_rows(values_only=True):
            if any(cell for cell in row):
                break
            first_row += 1

        for col in ws.iter_cols(values_only=True):
            if any(cell for cell in col):
                break
            first_col += 1

        return first_row, first_col

    @staticmethod
    def _get_xls_first_not_blank_row_and_col(ws):
        """
        功能：获取xls 非空白首行、首列索引
        """
        first_row = 0
        first_col = 0
        for i in range(ws.nrows):
            if any(cell for cell in ws.row_values(i)):
                first_row = i
                break

        for i in range(ws.ncols):
            if any(cell for cell in ws.col_values(i)):
                first_col = i
                break

        return first_row, first_col

    def lazy_load(self) -> Iterator[Document]:
        """
        ：返回：逐行读取表,返回 string list
        """
        file_check.SecFileCheck(self.file_path, self.MAX_SIZE).check()
        # 判断文件种类：支持 xlsx 与 xls 格式
        file_format = xlrd.inspect_format(self.file_path)
        if not file_format or (file_format != "xlsx" and file_format != 'xls'):
            raise ValueError('file type is not xlsx and xls')
        if self.file_path.endswith(XLRD_EXTENSION):
            return self._load_xls()
        elif self.file_path.endswith(OPENPYXL_EXTENSION):
            if self._is_zip_bomb():
                raise ValueError(f"file is a risk of zip bombs")
            else:
                return self._load_xlsx()
        else:
            raise TypeError(f"'{self.file_path}' file type is not one of (xlsx, xls).")

    def _get_xlsx_title(self, ws, first_row, first_col):
        title = []
        for col in range(first_col, ws.max_column + 1):
            ti = self._parse_xlsx_cell(ws, first_row, col)
            title.append(str(ti) if ti is not None else "")

        return title

    def _get_xls_title(self, ws, first_row, first_col):
        title = []
        for col in range(first_col, ws.ncols):
            ti = str(self._parse_xls_cell(ws, first_row, col))
            title.append(str(ti) if ti is not None else "")

        return title

    def _load_xlsx_one_sheet(self, ws, sheet_name):
        """
        功能：读取一个xlsx表单的值，每行值以title:value;title:value....的格式
        """
        lines = []
        if ws.max_row > MAX_ROW_NUM:
            logger.error(f"Exceeded maximum row limit of {MAX_ROW_NUM} in sheet '{sheet_name}'"
                         f" of file '{self.file_path}': {ws.max_row} rows found.")
            return lines
        if ws.max_column > MAX_COL_NUM:
            logger.error(f"Exceeded maximum row limit of {MAX_COL_NUM} in sheet '{sheet_name}'"
                         f" of file '{self.file_path}': {ws.max_column} rows found.")
            return lines

        blank_rows, blank_cols = self._get_xlsx_blank_rows_and_cols(ws)

        # 获取有效第一行,列的索引
        first_row, first_col = self._get_xlsx_first_not_blank_row_and_col(ws)

        # 判断表单是否有标题+内容，默认至少两行有效行
        if ws.max_row - len(blank_rows.keys()) < 2:
            return lines

        # 获取标题列表
        title = self._get_xlsx_title(ws, first_row, first_col)

        column_end = ws.max_column + 1
        for row_index in range(first_row + 1, ws.max_row + 1):
            # 空行无数据，不解析
            if row_index in blank_rows.keys():
                continue

            text_line = ""
            for col_index in range(1, column_end):
                # 空列无数据，不解析
                if col_index in blank_cols.keys():
                    continue

                val = self._parse_xlsx_cell(ws, row_index, col_index)
                ti = title[col_index - first_col]
                if not val:
                    continue

                text_line += str(ti).strip() + ":" + str(val).strip() + ";"

            lines.append(text_line)

        return lines

    def _load_xls_one_sheet(self, ws):
        """
        功能：读取一个xls表单的值，每行值以title:value;title:value....的格式
        """
        lines = []
        if ws.nrows > MAX_ROW_NUM:
            logger.error(f"Exceeded maximum row limit of {MAX_ROW_NUM} in sheet '{ws.name}'"
                         f" of file '{self.file_path}': {ws.nrows} rows found.")
            return lines
        if ws.ncols > MAX_COL_NUM:
            logger.error(f"Exceeded maximum row limit of {MAX_COL_NUM} in sheet '{ws.name}'"
                         f" of file '{self.file_path}': {ws.ncols} rows found.")
            return lines

        blank_rows, blank_cols = self._get_xls_blank_rows_and_cols(ws)
        # 获取有效第一行,列的索引
        first_row, first_col = self._get_xls_first_not_blank_row_and_col(ws)

        # 判断表单是否有标题+内容，默认至少两行有效行
        if ws.nrows - len(blank_rows.keys()) < 2:
            logger.info(f"In file ['{self.file_path}'], sheet ['{ws.name}'],"
                        f" not enough valid rows (at least two rows required). ")
            return lines

        # 获取标题列表
        title = self._get_xls_title(ws, first_row, first_col)
        ncols = ws.ncols
        for row_index in range(first_row + 1, ws.nrows):
            # 空行无数据，不解析
            if row_index in blank_rows.keys():
                continue

            text_line = ""
            for col_index in range(ncols):
                # 空列无数据，不解析
                if col_index in blank_cols.keys():
                    continue
                val = self._parse_xls_cell(ws, row_index, col_index)
                if not val:
                    continue

                ti = title[col_index - first_col]

                if ti in ["time"] and 0 <= float(val) <= 1:
                    text_line += str(ti) + ":" + str(self._exceltime_to_datetime(float(val))) + ";"
                else:
                    text_line += str(ti).strip() + ":" + str(val).strip() + ";"

            lines.append(text_line)

        return lines

    def _load_xls(self):
        wb = None
        try:
            wb = xlrd.open_workbook(self.file_path, formatting_info=True)
            if wb.nsheets > self.MAX_PAGE_NUM:
                logger.error(f"file '{self.file_path}' sheets number more than limit")
                return
            for i in range(wb.nsheets):
                ws = wb.sheet_by_index(i)
                lines = self._load_xls_one_sheet(ws)

                if not lines:
                    logger.info(f"In file ['{self.file_path}'] sheet ['{ws.name}'] is empty")
                    continue
                for line in lines:
                    yield Document(page_content=line,
                                   metadata={"source": self.file_path, "sheet": ws.name, "type": "text"})
        except xlrd.biffh.XLRDError as e:
            logger.error(f"Excel parsing error for file '{self.file_path}': {e}")
            return
        except IndexError as e:
            logger.error(f"Sheet index error in file '{self.file_path}': {str(e)}")
            return
        except Exception as e:
            logger.error(f"An error occurred while loading file '{self.file_path}': {e}")
            return
        finally:
            if wb:
                wb.release_resources()
                logger.info(f"file '{self.file_path}' Loading completed")

    def _load_xlsx(self):
        wb = None
        try:
            wb = load_workbook(self.file_path, data_only=True, keep_links=False)
            if len(wb.sheetnames) > self.MAX_PAGE_NUM:
                logger.error(f"file '{self.file_path}' sheets number more than limit")
                return
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines = self._load_xlsx_one_sheet(ws, sheet_name)
                if not lines:
                    logger.info(f"In file ['{self.file_path}'] sheet ['{sheet_name}'] is empty")
                    continue
                for line in lines:
                    yield Document(page_content=line, metadata={"source": self.file_path, "sheet": sheet_name})
        except KeyError as e:
            logger.error(f"Sheet not found in file '{self.file_path}': {str(e)}")
            return
        except Exception as e:
            logger.error(f"An error occurred while loading file '{self.file_path}': {e}")
            return
        finally:
            if wb:
                wb.close()
                logger.info(f"file '{self.file_path}' Loading completed")
